from __future__ import annotations

import io
import os
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from formulas import ExcelModel
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = APP_DIR / "PRESBYOPIA IOL SOFTWARE v2.xlsx"

SECTION_CONFIG = [
    {"title": "MONOFOCAL LENSES", "header_row": 4, "rows": [5, 6, 8, 9, 10, 11, 12], "cols": ["B", "C", "D"]},
    {"title": "EXTENDED MONOFOCAL and EDOF LENSES", "header_row": 14, "rows": [15, 16, 18, 19, 20, 21, 22], "cols": ["B", "C", "D"]},
    {"title": "DIFFRACTIVE LENSES", "header_row": 24, "rows": [25, 26, 28, 29, 30, 31, 32], "cols": ["B", "C", "D"]},
]

INPUT_CELLS = {
    "AGE": "A2",
    "PHOTOPIC PUPIL": "C2",
    "SCOTOPIC PUPIL": "D2",
    "ANGLE ALPHA": "E2",
    "ANGLE KAPPA": "F2",
    "CORNEAL SPHERICAL ABERRATION 6mm (Z40)": "G2",
    "CORNEAL VERTICAL COMA 6mm": "H2",
    "CORNEAL HORIZONTAL COMA 6mm": "I2",
    "CORNEAL ENDOTHELIAL DISEASE": "K2",
    "DRY EYE DISEASE": "K3",
    "DIABETIC RETINOPATHY": "K4",
    "MACULAR PATHOLOGY": "K5",
    "GLAUCOMA": "K6",
}


def load_template_bytes() -> bytes:
    return TEMPLATE_PATH.read_bytes()


def get_template_metadata() -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(load_template_bytes()), data_only=False)
    ws = wb["INPUT DATA"]
    number_formats: dict[str, str] = {}
    labels: dict[str, str] = {}
    headers: dict[int, list[str]] = {}
    for section in SECTION_CONFIG:
        header_row = section["header_row"]
        headers[header_row] = [ws[f"{col}{header_row}"].value for col in section["cols"]]
        for row in section["rows"]:
            labels[f"A{row}"] = ws[f"A{row}"].value
            for col in section["cols"]:
                coord = f"{col}{row}"
                number_formats[coord] = ws[coord].number_format
    defaults = {label: ws[cell].value for label, cell in INPUT_CELLS.items()}
    return {"number_formats": number_formats, "labels": labels, "headers": headers, "defaults": defaults}


def excel_display(value: Any, number_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if number_format == "0%":
            return f"{round(value * 100):.0f}%"
        if number_format == "0.00":
            return f"{value:.2f}"
        if number_format == "0.0":
            return f"{value:.1f}"
        if number_format == "0":
            return f"{value:.0f}"
        text = f"{value:.2f}"
        return text.rstrip("0").rstrip(".")
    return str(value)


def _extract_solution_value(solution: dict[str, Any], workbook_name: str, sheet_name: str, coord: str) -> Any:
    key = f"'[{workbook_name}]{sheet_name}'!{coord}"
    rng = solution[key]
    value = rng.value
    try:
        return value.item()
    except Exception:
        try:
            return value[0][0]
        except Exception:
            return value


def calculate_outputs(inputs: dict[str, Any]) -> dict[str, Any]:
    template_bytes = load_template_bytes()
    workbook = load_workbook(io.BytesIO(template_bytes), data_only=False)
    sheet = workbook["INPUT DATA"]
    for label, cell in INPUT_CELLS.items():
        sheet[cell] = inputs[label]

    metadata = get_template_metadata()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        workbook.save(tmp_path)
        workbook_name = os.path.basename(tmp_path)
        model = ExcelModel().loads(str(tmp_path)).finish()
        solution = model.calculate()
        section_tables: list[dict[str, Any]] = []
        for section in SECTION_CONFIG:
            header_row = section["header_row"]
            headers = metadata["headers"][header_row]
            records = []
            for row in section["rows"]:
                record: dict[str, Any] = {"PARAMETER": metadata["labels"][f"A{row}"]}
                for col, lens_name in zip(section["cols"], headers):
                    coord = f"{col}{row}"
                    raw_value = _extract_solution_value(solution, workbook_name, "INPUT DATA", coord)
                    record[lens_name] = excel_display(raw_value, metadata["number_formats"].get(coord, "General"))
                records.append(record)
            section_tables.append({"title": section["title"], "data": pd.DataFrame(records)})
        return {"tables": section_tables}
    finally:
        tmp_path.unlink(missing_ok=True)
